# @Author  :  comma 
# @Date    :  2020-11-22 11:21

'''
向Excel文件中写入数据
    创建工作簿对象     openpyxl.Workbook()
    获取活动工作表对象   wb.active
    获取单元格       sheet[单元格名称]
    向单元格中写入数据   cell.value=值
    向Excel中写入一行数据   sheet.append(列表)
    保存Excel文件   wb.save(文件)

从Excel中读取数据
    加载工作簿对象     openpyxl.load_workbook(文件名)
    获取活动工作表对象   wb.active
    获取单元格       sheet[单元格名称]
    获取单元格的值     cell.value
    获取一系列格子     sheet['A']，sheet['3'],sheet['A:C']
    获取整个表的所有行       sheet.rows
'''
import openpyxl

#
# # 创建工作簿对象
# wb = openpyxl.Workbook()
# # 获取工作表sheet
# sheet = wb.active
# # 获取指定的单元格
# cell = sheet['A1']
# # 向单元格写入数据
# cell.value = 'python学习'
# # 写入一行数据（列表）
# lst = ['comma', 'love', 'study']
# sheet.append(lst)
# # 添加多行数据
# lst2 = [['憨憨', '喜欢', '吃水果'], ['二豆', '喜欢吃', '番石榴'], ['渣渣', '喜欢吃', '榴莲']]
# for row in lst2:
#     sheet.append(row)
# # 保存
# wb.save('excel学习.xlsx')

############# 从Excel中读取数据 #################
# 加载Excel文件（创建一个Python中的工作簿对象）
lwb = openpyxl.load_workbook('excel学习.xlsx')
# 获取工作表对象
# sheet2=lwb.active   #只有一个表的时候
sheet2 = lwb['Sheet']  # 有多个表的时候可以根据名称来获取表
# 获取指定的单元格
cell1 = sheet2['A1']
value = cell1.value
print(value)
# 获取列数据
print('----------------------- 获取列数据----------------------------')
columns=sheet2['A']     # 获取A列的所有单元格
for col in columns:
    print(col.value)

# 获取行数据
print('----------------------- 获取行数据----------------------------')
row=sheet2[3]
for cell in row:
    print(cell.value)

# 获取多列数据
print('----------------------- 获取多列数据----------------------------')
cols=sheet2['A:C']
for col in cols:
    for cell in col:
        print(cell.value)
    print('--------每一列隔断-----------')